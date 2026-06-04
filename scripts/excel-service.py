#!/usr/bin/env python3
"""
Spiritual AI — Local Excel Sheet Service
Manages an Excel file that mirrors the Google Sheets schema.
Allows dynamic CRUD operations so user data (MBTI changes, sessions, etc.)
are reflected in real-time in the Excel file.
"""

import os
import json
import sys
import math
from datetime import datetime, date
from pathlib import Path
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import threading

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_DIR = Path.home() / ".spiritual-ai"
EXCEL_DIR.mkdir(exist_ok=True)
EXCEL_FILE = EXCEL_DIR / "spiritual_ai_data.xlsx"

# ── Sheet column definitions ──

USERS_COLUMNS = [
    "user_id", "first_seen", "last_seen", "email", "name",
    "country", "city", "device_type", "browser", "referral_source",
    "referral_csn", "entry_tag", "total_sessions", "total_reports",
    "total_revenue", "subscription_active", "opted_in_email",
]

SESSIONS_COLUMNS = [
    "session_id", "user_id", "started_at", "ended_at", "duration_seconds",
    "exchange_count", "drop_off_exchange", "completion_status",
    "entry_emotion_tag", "detected_emotions", "dominant_emotion",
    "confidence_score_final", "trigger_reason", "input_mode",
    "report_generated", "email_captured_at_exchange", "raw_transcript_url",
]

REPORTS_COLUMNS = [
    "csn", "sequence_number", "user_id", "session_id", "generated_at",
    "mbti_type", "mbti_confidence", "consciousness_identity",
    "archetype_symbol", "core_pattern", "jungian_complex", "root_belief",
    "root_age", "secondary_gain", "spiritual_path",
    "hawkins_level_current", "hawkins_level_target", "urgency_percent",
    "dharma_phase", "dob", "birth_time_approx", "birth_place",
    "lagna_sign", "moon_sign", "current_dasha", "witness_level",
    "report_url", "times_shared", "referrals_generated",
]

PRODUCTS_COLUMNS = [
    "transaction_id", "user_id", "csn", "purchased_at",
    "product_tier", "product_name", "architecture_matched",
    "amount_usd", "amount_local", "currency", "payment_method",
    "hours_after_report", "conversion_trigger", "subscription_status",
    "subscription_months", "mbti_of_buyer", "refund_issued", "refund_reason",
]

EVENTS_COLUMNS = [
    "timestamp", "user_id", "session_id", "event_type",
    "exchange_number", "event_value", "confidence_at_event", "emotion_at_event",
]

ALL_SHEETS = {
    "Users": USERS_COLUMNS,
    "Sessions": SESSIONS_COLUMNS,
    "Reports": REPORTS_COLUMNS,
    "Products": PRODUCTS_COLUMNS,
    "Events": EVENTS_COLUMNS,
}

# Header styling
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILLS = {
    "Users": PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid"),
    "Sessions": PatternFill(start_color="16213E", end_color="16213E", fill_type="solid"),
    "Reports": PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid"),
    "Products": PatternFill(start_color="533483", end_color="533483", fill_type="solid"),
    "Events": PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid"),
}
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def init_excel():
    """Create or open the Excel file with all sheet tabs and headers."""
    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for sheet_name, columns in ALL_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
        else:
            ws = wb[sheet_name]

        # Write headers if first row is empty or doesn't match
        header_ok = False
        if ws.max_row >= 1:
            first_row_vals = []
            for cell in ws[1]:
                first_row_vals.append(cell.value)
            header_ok = (first_row_vals[:len(columns)] == columns)

        if not header_ok:
            # Clear first row and rewrite headers
            for j, col in enumerate(columns):
                cell = ws.cell(row=1, column=j + 1, value=col)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILLS[sheet_name]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN_BORDER
            ws.row_dimensions[1].height = 20
            # Set column widths
            for j in range(len(columns)):
                col_letter = chr(65 + j) if j < 26 else chr(64 + j // 26) + chr(65 + j % 26)
                ws.column_dimensions[col_letter].width = min(max(len(columns[j]) + 4, 12), 40)
        # freeze panes at header
        ws.freeze_panes = "A2"

    wb.save(EXCEL_FILE)
    return wb


def ts_to_excel(dt) -> str:
    """Convert datetime to string for Excel."""
    if dt is None:
        return ""
    if isinstance(dt, (datetime, date)):
        return dt.isoformat()
    return str(dt)


def to_excel_value(val):
    """Convert Python values to Excel-safe values."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, float):
        if math.isnan(val):
            return 0
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    return val


class SheetManager:
    """Manages CRUD operations on the Excel file."""

    def __init__(self):
        init_excel()
        self._lock = threading.Lock()

    def _open(self):
        return load_workbook(EXCEL_FILE)

    def _save(self, wb):
        wb.save(EXCEL_FILE)

    def _find_sheet(self, wb, sheet_type):
        name_map = {
            "users": "Users",
            "sessions": "Sessions",
            "reports": "Reports",
            "products": "Products",
            "events": "Events",
        }
        name = name_map.get(sheet_type, sheet_type)
        return wb[name]

    def get_pk_col(self, sheet_type):
        pk_map = {
            "users": "user_id",
            "sessions": "session_id",
            "reports": "csn",
            "products": "transaction_id",
            "events": None,
            # events append-only
        }
        return pk_map.get(sheet_type)

    # ── Users ──

    def upsert_user(self, data: dict) -> dict:
        """Create or update a user row."""
        with self._lock:
            wb = self._open()
            ws = self._find_sheet(wb, "users")
            columns = USERS_COLUMNS

            # Find existing row by user_id
            user_id = data.get("user_id", "")
            row_idx = None
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, 1).value == user_id:
                    row_idx = r
                    break

            now = datetime.now().isoformat()
            row_data = []
            for col in columns:
                if col == "user_id":
                    row_data.append(to_excel_value(user_id))
                elif col == "first_seen" and row_idx is None:
                    row_data.append(to_excel_value(now))
                elif col == "first_seen" and row_idx is not None:
                    row_data.append(ws.cell(row_idx, columns.index(col) + 1).value or now)
                elif col == "last_seen":
                    row_data.append(to_excel_value(now))
                elif col in data:
                    row_data.append(to_excel_value(data[col]))
                elif row_idx is not None:
                    row_data.append(ws.cell(row_idx, columns.index(col) + 1).value)
                else:
                    row_data.append("")

            if row_idx is None:
                ws.append(row_data)
                action = "created"
            else:
                for j, val in enumerate(row_data):
                    ws.cell(row_idx, j + 1, val)
                action = "updated"

            self._save(wb)
            return {"action": action, "user_id": user_id}

    # ── Sessions ──

    def upsert_session(self, data: dict) -> dict:
        with self._lock:
            wb = self._open()
            ws = self._find_sheet(wb, "sessions")
            columns = SESSIONS_COLUMNS

            session_id = data.get("session_id", "")
            row_idx = None
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, 1).value == session_id:
                    row_idx = r
                    break

            row_data = []
            for col in columns:
                if col in data:
                    row_data.append(to_excel_value(data[col]))
                elif row_idx is not None:
                    row_data.append(ws.cell(row_idx, columns.index(col) + 1).value)
                else:
                    row_data.append(to_excel_value(self._session_default(col)))

            if row_idx is None:
                ws.append(row_data)
                action = "created"
            else:
                for j, val in enumerate(row_data):
                    ws.cell(row_idx, j + 1, val)
                action = "updated"

            self._save(wb)
            return {"action": action, "session_id": session_id}

    def _session_default(self, col):
        defaults = {
            "duration_seconds": 0,
            "exchange_count": 0,
            "completion_status": "in_progress",
            "input_mode": "text",
            "report_generated": "FALSE",
        }
        return defaults.get(col, "")

    # ── Reports ──

    def upsert_report(self, data: dict) -> dict:
        with self._lock:
            wb = self._open()
            ws = self._find_sheet(wb, "reports")
            columns = REPORTS_COLUMNS

            csn = data.get("csn", "")
            row_idx = None
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, 1).value == csn:
                    row_idx = r
                    break

            row_data = []
            for col in columns:
                if col in data:
                    row_data.append(to_excel_value(data[col]))
                elif row_idx is not None:
                    row_data.append(ws.cell(row_idx, columns.index(col) + 1).value)
                else:
                    row_data.append("")

            if row_idx is None:
                ws.append(row_data)
                action = "created"
            else:
                for j, val in enumerate(row_data):
                    ws.cell(row_idx, j + 1, val)
                action = "updated"

            self._save(wb)
            return {"action": action, "csn": csn}

    # ── Products ──

    def append_product(self, data: dict) -> dict:
        with self._lock:
            wb = self._open()
            ws = self._find_sheet(wb, "products")
            columns = PRODUCTS_COLUMNS

            row_data = []
            for col in columns:
                row_data.append(to_excel_value(data.get(col, "")))

            ws.append(row_data)
            self._save(wb)
            return {"action": "created", "transaction_id": data.get("transaction_id", "")}

    # ── Events ──

    def append_event(self, data: dict) -> dict:
        with self._lock:
            wb = self._open()
            ws = self._find_sheet(wb, "events")
            columns = EVENTS_COLUMNS

            row_data = []
            for col in columns:
                if col == "timestamp" and "timestamp" not in data:
                    row_data.append(to_excel_value(datetime.now().isoformat()))
                else:
                    row_data.append(to_excel_value(data.get(col, "")))

            ws.append(row_data)
            self._save(wb)
            return {"action": "appended"}

    # ── Read operations ──

    def get_all(self, sheet_type: str) -> list:
        with self._lock:
            wb = self._open()
            ws = self._find_sheet(wb, sheet_type)
            columns = ALL_SHEETS[ws.title]
            results = []
            for r in range(2, ws.max_row + 1):
                row = {}
                has_val = False
                for j, col in enumerate(columns):
                    val = ws.cell(r, j + 1).value
                    if val is not None and val != "":
                        has_val = True
                    row[col] = val
                if has_val:
                    results.append(row)
            return results

    def get_user(self, user_id: str) -> dict | None:
        with self._lock:
            wb = self._open()
            ws = self._find_sheet(wb, "users")
            columns = USERS_COLUMNS
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, 1).value == user_id:
                    row = {}
                    for j, col in enumerate(columns):
                        row[col] = ws.cell(r, j + 1).value
                    return row
            return None

    def get_summary(self) -> dict:
        with self._lock:
            wb = self._open()
            summary = {}
            for sheet_name in ALL_SHEETS:
                ws = wb[sheet_name]
                summary[sheet_name] = {
                    "rows": ws.max_row - 1,
                    "columns": len(ALL_SHEETS[sheet_name]),
                }
            return summary


# ── HTTP Server ──

manager = SheetManager()


class SheetHTTPHandler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        # Suppress default logging noise
        pass

    def _send_json(self, data, status=200):
        body = json.dumps(data, indent=2, default=str).encode()
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path.rstrip("/")

        if path == "/health":
            self._send_json({"status": "ok", "file": str(EXCEL_FILE)})
            return

        if path == "/summary":
            self._send_json(manager.get_summary())
            return

        # /sheet/{type} — get all rows
        if path.startswith("/sheet/"):
            sheet_type = path.split("/")[-1]
            if sheet_type not in ("users", "sessions", "reports", "products", "events"):
                self._send_json({"error": "Unknown sheet type"}, 400)
                return
            data = manager.get_all(sheet_type)
            self._send_json({"sheet": sheet_type, "data": data, "count": len(data)})
            return

        # /user/{id}
        if path.startswith("/user/"):
            user_id = path.split("/")[-1]
            data = manager.get_user(user_id)
            if data:
                self._send_json(data)
            else:
                self._send_json({"error": "User not found"}, 404)
            return

        self._send_json({"error": "Not found"}, 404)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_OPTIONS(self):
        self.def_do_OPTIONS()

    def do_POST(self):
        content_length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(content_length) if content_length > 0 else b"{}"

        try:
            data = json.loads(body) if body else {}
        except json.JSONDecodeError:
            self._send_json({"error": "Invalid JSON"}, 400)
            return

        path = self.path.rstrip("/")

        if path == "/upsert" or path == "/upsert/user":
            result = manager.upsert_user(data)
            self._send_json(result)
            return

        if path == "/upsert/session":
            result = manager.upsert_session(data)
            self._send_json(result)
            return

        if path == "/upsert/report":
            result = manager.upsert_report(data)
            self._send_json(result)
            return

        if path == "/append/product":
            result = manager.append_product(data)
            self._send_json(result)
            return

        if path == "/append/event":
            result = manager.append_event(data)
            self._send_json(result)
            return

        # Generic: /upsert/{sheet_type}
        if path.startswith("/upsert/"):
            sheet_type = path.split("/")[-1]
            if sheet_type == "user":
                result = manager.upsert_user(data)
            elif sheet_type == "session":
                result = manager.upsert_session(data)
            elif sheet_type == "report":
                result = manager.upsert_report(data)
            else:
                self._send_json({"error": f"Unsupported sheet type: {sheet_type}"}, 400)
                return
            self._send_json(result)
            return

        self._send_json({"error": f"Unknown endpoint: {path}"}, 404)


def run_server(port=8765):
    server = HTTPServer(("127.0.0.1", port), SheetHTTPHandler)
    print(f"📊 Spiritual AI Excel Service running on http://127.0.0.1:{port}")
    print(f"📁 Excel file: {EXCEL_FILE}")
    init_excel()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nShutting down Excel service...")
        server.shutdown()


if __name__ == "__main__":
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8765
    run_server(port)
